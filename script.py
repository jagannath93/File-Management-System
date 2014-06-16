
# openpyxl imports
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
from openpyxl.cell import get_column_letter

from fms.models import *

def load_data():
  try:
    wb = load_workbook(filename = "RACKI.xlsx")
    ws = wb.get_sheet_by_name("Architecture")
    for row in ws.rows:
      rack = 'RACK I'
      category = 'Architecture'
      shelf_set = 'set-1'

      code = row[0].value
      name = row[1].value
      author =  row[2].value
      publisher = row[3].value
      isbn = row[4].value
      _status = row[5].value
      row = row[6].value

      status = True
      if _status == 'Y':
        status = True
      else:
        status = False
      
      tmp = Book.objects.filter(code=code)
      if len(tmp) is 0:
        address = BookAddress(rack=rack, row=row, shelf_set=shelf_set) 
        address.save()
        cat = Category.objects.get(name=category)
        book = Book(name=name, code=code, author=author, publisher=publisher, isbn_number=isbn, avilability_status=status, address=address)
        book.save()
        book.categories.add(cat)
        print code+": - DONE."
      else:
        print code+": - Already existing."
  except Exception as e:
    print e

