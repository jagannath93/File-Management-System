
# openpyxl imports
from openpyxl.reader.excel import load_workbook
#from openpyxl.writer.excel import excelwriter
from openpyxl.workbook import workbook
from openpyxl.cell import get_column_letter

from fms.models import *
import os.path

def load_data(arg1, arg2):
  try:
    wb = load_workbook(filename = "LIB-LIST.xlsx")
    ws = wb.get_sheet_by_name(arg1)
    for row in ws.rows:
      address = BookAddress.objects.get(rack=arg2)
      category = Category.objects.get(name=arg1)
      #shelf_set = 'set-1'

      code = row[0].value
      name = row[1].value
      author =  row[2].value
      publisher = row[3].value
      isbn = row[4].value
      _status = row[5].value
      #row = row[6].value

      status = True
      if _status == 'Y':
        status = True
      else:
        status = False
      
      tmp = Book.objects.filter(code=code)
      if len(tmp) is 0:
        #address = BookAddress(rack=rack, row=row, shelf_set=shelf_set) 
        #address.save()
        #cat = Category.objects.get(name=category)
        book = Book(name=name, code=code, author=author, publisher=publisher, isbn_number=isbn, avilability_status=status, address=address)
        book.save()
        book.categories.add(category)
        if os.path.isfile('media/books/'+arg1+'/'+book.code+'.jpg'): 
          image = File(open('media/books/'+arg1+'/'+book.code+'.jpg')) 
          book.image = image
          book.save()
        print code+": - DONE."
      else:
        tmp[0].name = name
        tmp[0].author = author
        tmp[0].publisher = publisher
        tmp[0].isbn_number = isbn
        tmp[0].avilability_status = status
        tmp[0].address = address
        if os.path.isfile('media/books/'+arg1+'/'+code+'.jpg'): 
          image = File(open('media/books/'+arg1+'/'+code+'.jpg')) 
          tmp[0].image = image
        tmp[0].save()
        print code+": - Already existing."
  except Exception as e:
    print e
