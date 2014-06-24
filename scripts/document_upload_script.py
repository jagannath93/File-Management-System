# Script to load all document data from an excel file to DATABASE.

# openpyxl imports
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import workbook
from openpyxl.cell import get_column_letter
import sys

from fms.models import *

def load_data(*args):
  try:
    wb = load_workbook(filename = "Doc.xlsx")
    #sheet_name = sys.argv[1]
    #sheet_name = raw_input("Enter Sheet Name(IN CAPS): ")
    #input_args = .split()

    ws = wb.get_sheet_by_name(args[0])
    for row in ws.rows:
      address = row[0].value
      name = row[1].value
      rack_name = row[2].value

      tmp = Document.objects.filter(name=name)
      if len(tmp) is 0:
        _ad = address.split('/')
        doc = None
        if len(_ad) is 4:
          cat = DocumentCategory.objects.get(code=_ad[0])
          subcat1 = DocumentSubCategory1.objects.get(code = _ad[1], cat=cat)
          subcat2 = DocumentSubCategory2.objects.get(code = _ad[2], cat=cat, subcat1=subcat1)
          doc = Document(name=name, cat=cat, subcat1=subcat1, subcat2=subcat2, document_number=_ad[3])
          doc.save()
        elif len(_ad) is 3:
          cat = DocumentCategory.objects.get(code=_ad[0])
          subcat1 = DocumentSubCategory1.objects.get(code = _ad[1], cat=cat)
          try:
            subcat2 = DocumentSubCategory2.objects.get(code = _ad[2], subcat1=subcat1, cat=cat)
            doc = Document(name=name, cat=cat, subcat1=subcat1, subcat2=subcat2)
          except:
            doc = Document(name=name, cat=cat, subcat1=subcat1, document_number=_ad[2])
            pass
          doc.save()
        elif len(_ad) is 2:
          cat = DocumentCategory.objects.get(code=_ad[0])
          try:
            subcat1 = DocumentSubCategory1.objects.get(code = _ad[1], cat=cat)
            doc = Document(name=name, cat=cat, subcat1=subcat1)
          except:
            doc = Document(name=name, cat=cat, document_number=_ad[1])
            pass
          doc.save()
        elif len(_ad) is 1:
          doc = Document(name=name, document_number=_ad[0])
          doc.save()

        try:
          rack = DocumentRack.objects.get(rack_name=rack_name)
          doc.rack = rack
          doc.save()
        except Exception as e:
          print e
        print address +":  OK"
      else:
        print address +":  Already existing..."
  except Exception as e:
    print e
     
def load_racks():
  try:
    wb = load_workbook(filename = "RACK.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")
    for row in ws.rows:
      name = row[0].value
      doc_type = row[1].value
      _type = row[2].value

      tmp = DocumentRack.objects.filter(rack_name=name)
      if len(tmp) is 0:
        doc = DocumentRack(rack_name=name, document_type=doc_type, type=_type)
        doc.save()
        print name+"  OK"
      else:
        print name+"  Already There..."
  except Exception as e:
    print e
